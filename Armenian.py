import telebot
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, KeyboardButton
from datetime import date, timedelta, datetime
from pathlib import Path
import json
import time
import re
import threading
from flask import Flask
import os
import openpyxl
from io import BytesIO

# ====== мини-вебсервер для Render ======
app = Flask(__name__)

@app.route("/")
def home():
    return "Attendance bot is alive!"

def run_flask():
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)

threading.Thread(target=run_flask, daemon=True).start()
# =======================================

# ================== НАСТРОЙКИ ==================
TOKEN = "8300421705:AAGzPJIF1Ia-GjHK-lriCeuDzOd39PgdvOg"
bot = telebot.TeleBot(TOKEN)

# Удаляем вебхук на всякий случай
try:
    bot.remove_webhook()
except Exception as e:
    print("Ошибка при удалении webhook:", e)

MAIN_ADMIN_ID = 1509389908  # Ваш ID (староста)
ADMIN_IDS = {6294378266, 5460930562, 5180067949, 6383870956, MAIN_ADMIN_ID}  # Множество админов

REFERENCE_MONDAY = date(2026, 1, 12)
REFERENCE_WEEK_TYPE = "чисельник"

# Файлы для хранения данных
USERS_FILE = "users.json"           # информация о пользователях (id, группа, фио, role)
STUDENTS_FILE = "students.json"      # список студентов с привязкой к группе (ФИО, tg_id, группа)
ATTENDANCE_FILE = "attendance.json"  # журнал посещаемости
SCHEDULE_FILE = "schedule.json"      # расписание
SUBSTITUTIONS_FILE = "substitutions.json"  # замены на конкретные даты
TEMP_CHANGES_FILE = "temp_changes.json"    # временные изменения (до конца недели)
CHANGELOG_FILE = "changelog.json"          # журнал изменений

# Расписание звонков
BELL_SCHEDULE = {
    "monday": {
        1: "08:30–09:50",
        2: "10:00–11:20",
        3: "11:50–13:10",
        4: "14:00–15:20",
        5: "15:30–16:50",
    },
    "other": {
        1: "08:30–09:50",
        2: "10:00–11:20",
        3: "11:50–13:10",
        4: "13:20–14:40",
        5: "14:50–16:10",
    },
}

DAY_ALIASES = {
    "понеділок": "monday", "понедельник": "monday", "пн": "monday", "пн.": "monday", "пон": "monday", "пон.": "monday", "mon": "monday", "monday": "monday",
    "вівторок": "tuesday", "вторник": "tuesday", "вт": "tuesday", "вт.": "tuesday", "втор": "tuesday", "tue": "tuesday", "tuesday": "tuesday",
    "середа": "wednesday", "середу": "wednesday", "ср": "wednesday", "ср.": "wednesday", "среда": "wednesday", "среду": "wednesday", "wed": "wednesday", "wednesday": "wednesday",
    "четвер": "thursday", "четверг": "thursday", "чт": "thursday", "чт.": "thursday", "чтв": "thursday", "thu": "thursday", "thursday": "thursday",
    "пʼятниця": "friday", "п'ятниця": "friday", "пʼятницю": "friday", "п'ятницю": "friday", "пятница": "friday", "пятницу": "friday", "пт": "friday", "пт.": "friday", "пят": "friday", "fri": "friday", "friday": "friday",
    "субота": "saturday", "суботу": "saturday", "суббота": "saturday", "субботу": "saturday", "сб": "saturday", "сб.": "saturday", "sat": "saturday", "saturday": "saturday",
    "неділя": "sunday", "неділю": "sunday", "воскресенье": "sunday", "неделя": "sunday", "нд": "sunday", "нд.": "sunday", "вс": "sunday", "вс.": "sunday", "вск": "sunday", "sun": "sunday", "sunday": "sunday",
}

DAYS_RU = {
    "monday": "Понеділок",
    "tuesday": "Вівторок",
    "wednesday": "Середа",
    "thursday": "Четвер",
    "friday": "Пʼятниця",
    "saturday": "Субота",
    "sunday": "Неділя",
}

NO_LESSON_SUBJECTS = {
    "немає пари", "нема пари", "нет пары", "немає уроку", "нема уроку", 
    "уроку немає", "-", "", " ",
}

# ================== РАСПИСАНИЕ ==================
def create_schedule_bcig():
    return {
        "monday": {
            "чисельник": {
                "1": {"subject": "Фізика і астрономія",  "room": "129", "teacher": "Гуленко І.А."},
                "2": {"subject": "Українська література","room": "115", "teacher": "Лосєва К.С."},
                "3": {"subject": "Історія України",      "room": "114", "teacher": "Мелещук Ю.Л."},
                "org": {"subject": "Організаційна година","room": "205", "teacher": "Крамаренко Л.О."},
                "4": {"subject": "Зарубіжна література", "room": "116", "teacher": "Мещерякова О.В."},
            },
            "знаменник": {
                "1": {"subject": "Фізика і астрономія",  "room": "129", "teacher": "Гуленко І.А."},
                "2": {"subject": "Українська література","room": "115", "teacher": "Лосєва К.С."},
                "3": {"subject": "Всесвітня історія",    "room": "114", "teacher": "Мелещук Ю.Л."},
                "org": {"subject": "Організаційна година","room": "205", "teacher": "Крамаренко Л.О."},
            },
        },
        "tuesday": {
            "чисельник": {
                "2": {"subject": "Фізична культура", "room": "с/з №2", "teacher": "Багрін В.С."},
                "3": {"subject": "Математика",       "room": "121",  "teacher": "Приймак О.В."},
                "4": {"subject": "Українська мова",  "room": "307",  "teacher": "Гавриленко С.Т."},
            },
            "знаменник": {
                "2": {"subject": "Фізична культура", "room": "с/з №2", "teacher": "Багрін В.С."},
                "3": {"subject": "Математика",       "room": "121",  "teacher": "Приймак О.В."},
                "4": {"subject": "Українська мова",  "room": "307",  "teacher": "Гавриленко С.Т."},
            },
        },
        "wednesday": {
            "чисельник": {
                "2": {"subject": "Технології",          "room": "208", "teacher": "Потапова А.О."},
                "3": {"subject": "Біологія і екологія", "room": "16",  "teacher": "Золотова К.В."},
                "4": {"subject": "Захист України",      "room": "242 / 201", "teacher": "Санко / Киянчук"},
            },
            "знаменник": {
                "2": {"subject": "Математика",          "room": "121", "teacher": "Приймак О.В."},
                "3": {"subject": "Біологія і екологія", "room": "16",  "teacher": "Золотова К.В."},
                "4": {"subject": "Захист України",      "room": "242 / 201", "teacher": "Санко / Киянчук"},
            },
        },
        "thursday": {
            "чисельник": {
                "1": {"subject": "Історія України",     "room": "114", "teacher": "Мелещук Ю.Л."},
                "2": {"subject": "Громадянська освіта", "room": "142", "teacher": "Зубко Г.М."},
                "3": {"subject": "Інформатика",         "room": "39",  "teacher": "Короленко / Єреп"},
            },
            "знаменник": {
                "1": {"subject": "Історія України",     "room": "114", "teacher": "Мелещук Ю.Л."},
                "2": {"subject": "Громадянська освіта", "room": "142", "teacher": "Зубко Г.М."},
                "3": {"subject": "Інформатика",         "room": "39",  "teacher": "Короленко / Єреп"},
                "4": {"subject": "Географія",           "room": "123", "teacher": "Баранець Т.О."},
            },
        },
        "friday": {
            "чисельник": {
                "2": {"subject": "Іноземна мова",       "room": "224 a", "teacher": "Криваноченкова Л.І."},
                "3": {"subject": "Хімія",               "room": "16",    "teacher": "Золотова К.В."},
                "4": {"subject": "Фізика і астрономія", "room": "129",   "teacher": "Гуленко І.А."},
            },
            "знаменник": {
                "2": {"subject": "Іноземна мова",       "room": "224 a", "teacher": "Криваноченкова Л.І."},
                "3": {"subject": "Хімія",               "room": "16",    "teacher": "Золотова К.В."},
                "4": {"subject": "Фізична культура",    "room": "с/з №2", "teacher": "Багрін В.С."},
            },
        },
        "saturday": {},
        "sunday":   {},
    }

def create_schedule_bcis():
    return {
        "monday": {
            "чисельник": {
                "1": {"subject": "Фізична культура", "room": "с/з №2", "teacher": "Свиридов А.П."},
                "2": {"subject": "Іноземна мова",    "room": "224 а", "teacher": "Криваноченкова Л.І."},
                "3": {"subject": "Математика",       "room": "121",  "teacher": "Приймак О.В."},
                "org": {"subject": "Організаційна година","room": "205", "teacher": "Крамаренко Л.О."}
            },
            "знаменник": {
                "1": {"subject": "Фізична культура", "room": "с/з №2", "teacher": "Свиридов А.П."},
                "2": {"subject": "Іноземна мова",    "room": "224 а", "teacher": "Криваноченкова Л.І."},
                "3": {"subject": "Математика",       "room": "121",  "teacher": "Приймак О.В."},
                "org": {"subject": "Організаційна година","room": "205", "teacher": "Крамаренко Л.О."},
            },
        },
        "tuesday": {
            "чисельник": {
                "1": {"subject": "Біологія і екологія", "room": "16",  "teacher": "Золотова К.В."},
                "2": {"subject": "Історія України",     "room": "114", "teacher": "Меленчук Ю.Д."},
                "3": {"subject": "Інформатика",         "room": "39",  "teacher": "Короленко / Єреп"},
            },
            "знаменник": {
                "1": {"subject": "Біологія і екологія", "room": "16",  "teacher": "Золотова К.В."},
                "2": {"subject": "Всесвітня історія",   "room": "114", "teacher": "Меленчук Ю.Д."},
                "3": {"subject": "Інформатика",         "room": "39",  "teacher": "Короленко / Єреп"},
            },
        },
        "wednesday": {
            "чисельник": {
                "1": {"subject": "Хімія",               "room": "16",  "teacher": "Золотова К.В."},
                "2": {"subject": "Математика",          "room": "121", "teacher": "Приймак О.В."},
                "3": {"subject": "Захист України",      "room": "242 / 201", "teacher": "Санко / Киянчук"},
            },
            "знаменник": {
                "1": {"subject": "Хімія",               "room": "16",  "teacher": "Золотова К.В."},
                "2": {"subject": "Технології",          "room": "208", "teacher": "Потапова А.О."},
                "3": {"subject": "Захист України",      "room": "242 / 201", "teacher": "Санко / Киянчук"},
                "4": {"subject": "Фізика і астрономія", "room": "129", "teacher": "Гуленко І.А."},
            },
        },
        "thursday": {
            "чисельник": {
                "1": {"subject": "Громадянська освіта", "room": "142", "teacher": "Зубко Г.М."},
                "2": {"subject": "Фізика і астрономія", "room": "129", "teacher": "Гуленко І.А."},
                "3": {"subject": "Українська мова",     "room": "307", "teacher": "Гавриленко С.Т."},
                "4": {"subject": "Зарубіжна література","room": "116", "teacher": "Менцєрякова О.В."},
            },
            "знаменник": {
                "1": {"subject": "Громадянська освіта", "room": "142", "teacher": "Зубко Г.М."},
                "2": {"subject": "Фізика і астрономія", "room": "129", "teacher": "Гуленко І.А."},
                "3": {"subject": "Українська мова",     "room": "307", "teacher": "Гавриленко С.Т."},
            },
        },
        "friday": {
            "чисельник": {
                "1": {"subject": "Фізична культура", "room": "с/з №2", "teacher": "Свиридов А.П."},
                "2": {"subject": "Історія України",   "room": "114", "teacher": "Меленчук Ю.Д."},
                "3": {"subject": "Українська література","room": "115", "teacher": "Лосєва К.С."},
            },
            "знаменник": {
                "1": {"subject": "Географія",         "room": "123", "teacher": "Бараненко Т.О."},
                "2": {"subject": "Історія України",   "room": "114", "teacher": "Меленчук Ю.Д."},
                "3": {"subject": "Українська література","room": "115", "teacher": "Лосєва К.С."},
            },
        },
        "saturday": {},
        "sunday":   {},
    }

# Загружаем или создаём расписание
def load_schedule():
    path = Path(SCHEDULE_FILE)
    if not path.exists():
        return {
            "БЦІГ-25": create_schedule_bcig(),
            "БЦІСТ-25": create_schedule_bcis()
        }
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def save_schedule(data):
    path = Path(SCHEDULE_FILE)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

schedule = load_schedule()

# ================== ВРЕМЕННЫЕ ИЗМЕНЕНИЯ (ЗАМЕНЫ НА ДЕНЬ НЕДЕЛИ) ==================
def load_temp_changes():
    path = Path(TEMP_CHANGES_FILE)
    if not path.exists():
        return {"БЦІГ-25": {}, "БЦІСТ-25": {}}
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
        if "БЦІГ-25" not in data:
            data["БЦІГ-25"] = {}
        if "БЦІСТ-25" not in data:
            data["БЦІСТ-25"] = {}
        return data

def save_temp_changes():
    path = Path(TEMP_CHANGES_FILE)
    with path.open("w", encoding="utf-8") as f:
        json.dump(temp_changes, f, ensure_ascii=False, indent=2)

temp_changes = load_temp_changes()

def load_changelog():
    path = Path(CHANGELOG_FILE)
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def save_changelog():
    path = Path(CHANGELOG_FILE)
    with path.open("w", encoding="utf-8") as f:
        json.dump(changelog, f, ensure_ascii=False, indent=2)

changelog = load_changelog()

# ================== ЗАМЕНЫ НА КОНКРЕТНЫЕ ДАТЫ (старая система) ==================
def load_substitutions():
    path = Path(SUBSTITUTIONS_FILE)
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def save_substitutions():
    path = Path(SUBSTITUTIONS_FILE)
    with path.open("w", encoding="utf-8") as f:
        json.dump(substitutions, f, ensure_ascii=False, indent=2)

substitutions = load_substitutions()

# ================== ПОЛЬЗОВАТЕЛИ ==================
def load_users():
    path = Path(USERS_FILE)
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def save_users():
    path = Path(USERS_FILE)
    with path.open("w", encoding="utf-8") as f:
        json.dump(users, f, ensure_ascii=False, indent=2)

users = load_users()

# ================== СТУДЕНТЫ ==================
def load_students():
    path = Path(STUDENTS_FILE)
    if not path.exists():
        # При первом запуске создаём список студентов из твоего списка
        initial_students = [
            # БЦІСТ-25 с ID
            {"tg_id": 1271426468, "full_name": "Поліна Коняхіна", "group": "БЦІСТ-25", "username": ""},
            {"tg_id": 8308504155, "full_name": "Валерія Осадча", "group": "БЦІСТ-25", "username": ""},
            {"tg_id": 964331404,  "full_name": "Роман Шмельов", "group": "БЦІСТ-25", "username": ""},
            {"tg_id": 2072289628, "full_name": "Катерина Вініченко", "group": "БЦІСТ-25", "username": ""},
            {"tg_id": 5201471830, "full_name": "Аляб'єв Гліб", "group": "БЦІСТ-25", "username": ""},
            # БЦІСТ-25 без ID
            {"tg_id": None, "full_name": "Олександра Побережна", "group": "БЦІСТ-25", "username": ""},
            {"tg_id": None, "full_name": "Ліна Березіна", "group": "БЦІСТ-25", "username": ""},
            {"tg_id": None, "full_name": "Орина Карпа", "group": "БЦІСТ-25", "username": ""},
            {"tg_id": None, "full_name": "Ленюк Валерія", "group": "БЦІСТ-25", "username": ""},
            {"tg_id": None, "full_name": "Злата Бурцева", "group": "БЦІСТ-25", "username": ""},
            {"tg_id": None, "full_name": "Вікторія Палець", "group": "БЦІСТ-25", "username": ""},
            {"tg_id": None, "full_name": "Тимофій Іванов", "group": "БЦІСТ-25", "username": ""},
            {"tg_id": None, "full_name": "Марія Забєліна", "group": "БЦІСТ-25", "username": ""},
            {"tg_id": None, "full_name": "Ковпак Андрій", "group": "БЦІСТ-25", "username": ""},
            {"tg_id": None, "full_name": "Чернишов Артем", "group": "БЦІСТ-25", "username": ""},
            {"tg_id": None, "full_name": "Федоренко Захар", "group": "БЦІСТ-25", "username": ""},
            {"tg_id": None, "full_name": "Дмитренко Анастасія", "group": "БЦІСТ-25", "username": ""},
            {"tg_id": None, "full_name": "Боліла Валерія", "group": "БЦІСТ-25", "username": ""},
            # БЦІГ-25 с ID
            {"tg_id": 5460930562, "full_name": "Головко Олексій", "group": "БЦІГ-25", "username": ""},
            {"tg_id": 6690780079, "full_name": "Колодуб Богдана", "group": "БЦІГ-25", "username": ""},
            {"tg_id": 1649793559, "full_name": "Петренко Ярослава", "group": "БЦІГ-25", "username": ""},
            {"tg_id": 5542839738, "full_name": "Пасічник Марія", "group": "БЦІГ-25", "username": ""},
            {"tg_id": 6700437572, "full_name": "Лоцман Анна", "group": "БЦІГ-25", "username": ""},
            {"tg_id": 1045528149, "full_name": "Лабур Даніїл", "group": "БЦІГ-25", "username": ""},
            # БЦІГ-25 без ID
            {"tg_id": None, "full_name": "Горбань Євгеній", "group": "БЦІГ-25", "username": ""},
            {"tg_id": None, "full_name": "Беседін Роман", "group": "БЦІГ-25", "username": ""},
            {"tg_id": None, "full_name": "Прокопенко Дмитро", "group": "БЦІГ-25", "username": ""},
            {"tg_id": None, "full_name": "Песнєва Крістіна", "group": "БЦІГ-25", "username": ""},
            {"tg_id": None, "full_name": "Павленко Ольга", "group": "БЦІГ-25", "username": ""},
            {"tg_id": None, "full_name": "Мальчиков Єгор", "group": "БЦІГ-25", "username": ""},
            {"tg_id": None, "full_name": "Лупарев Артем", "group": "БЦІГ-25", "username": ""},
            {"tg_id": None, "full_name": "Іноземцев Ярослав", "group": "БЦІГ-25", "username": ""},
            {"tg_id": None, "full_name": "Янчук Олена", "group": "БЦІГ-25", "username": ""},
            {"tg_id": None, "full_name": "Хавер Єфим", "group": "БЦІГ-25", "username": ""},
        ]
        return initial_students
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def save_students():
    path = Path(STUDENTS_FILE)
    with path.open("w", encoding="utf-8") as f:
        json.dump(students, f, ensure_ascii=False, indent=2)

students = load_students()

# ================== ЖУРНАЛ ПОСЕЩАЕМОСТИ ==================
def load_attendance():
    path = Path(ATTENDANCE_FILE)
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)

def save_attendance():
    path = Path(ATTENDANCE_FILE)
    with path.open("w", encoding="utf-8") as f:
        json.dump(attendance, f, ensure_ascii=False, indent=2)

attendance = load_attendance()

# ================== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==================
def get_week_type(target_date=None):
    if target_date is None:
        target_date = date.today()
    delta_days = (target_date - REFERENCE_MONDAY).days
    weeks_passed = delta_days // 7
    if weeks_passed % 2 == 0:
        return REFERENCE_WEEK_TYPE
    else:
        return "знаменник" if REFERENCE_WEEK_TYPE == "чисельник" else "чисельник"

def get_day_key(target_date=None):
    if target_date is None:
        target_date = date.today()
    weekday = target_date.weekday()
    mapping = {0: "monday", 1: "tuesday", 2: "wednesday", 3: "thursday", 4: "friday", 5: "saturday", 6: "sunday"}
    return mapping[weekday]

def get_pair_time(day_key, pair_num):
    if day_key == "monday":
        return BELL_SCHEDULE["monday"].get(pair_num)
    else:
        return BELL_SCHEDULE["other"].get(pair_num)

def get_students_by_group(group):
    return [s for s in students if s["group"] == group]

def is_admin(user_id):
    return user_id in ADMIN_IDS

def remember_user(message):
    u = message.from_user
    uid = str(u.id)
    if uid not in users:
        users[uid] = {
            "id": u.id,
            "username": u.username or "",
            "first_name": u.first_name or "",
            "role": "student",
            "group": None,
            "full_name": None,
            "registered": False
        }
        save_users()

def get_schedule_with_changes(group_name, day_key, week_type):
    """Возвращает расписание для группы, дня и недели с учётом временных изменений"""
    if group_name not in schedule:
        return {}
    day_data = schedule[group_name].get(day_key, {})
    day_schedule = day_data.get(week_type, {}).copy()
    if (group_name in temp_changes and
        day_key in temp_changes[group_name] and
        week_type in temp_changes[group_name][day_key]):
        for pair_num, change in temp_changes[group_name][day_key][week_type].items():
            day_schedule[pair_num] = {
                "subject": change["subject"],
                "room": change.get("room", ""),
                "teacher": change.get("teacher", "")
            }
    return day_schedule

# ================== КОМАНДЫ ДЛЯ ВСЕХ ==================
@bot.message_handler(commands=["start", "help"])
def send_welcome(message):
    remember_user(message)
    uid = str(message.from_user.id)
    user = users[uid]

    # Якщо користувач вже зареєстрований
    if user.get("registered"):
        text = (
            f"Привіт, {user.get('full_name', '')}!\n"
            f"Твоя група: {user.get('group', 'не вказана')}\n\n"
            "Команди:\n"
            "/my_stats – моя статистика відвідування\n"
            "/mygroup – показати мою групу\n"
        )
        if is_admin(message.from_user.id):
            text += "\n👑 Адмін-команди:\n/adminhelp"
        bot.reply_to(message, text)
        return

    # Перевіряємо, чи є студент із таким tg_id у списку студентів
    student_entry = next((s for s in students if s.get("tg_id") == message.from_user.id), None)
    if student_entry:
        # Автоматично реєструємо
        users[uid]["full_name"] = student_entry["full_name"]
        users[uid]["group"] = student_entry["group"]
        users[uid]["registered"] = True
        save_users()
        text = (
            f"Привіт, {student_entry['full_name']}!\n"
            f"Твоя група: {student_entry['group']}\n\n"
            "Ти вже є в списку студентів. Тепер можеш користуватись ботом.\n"
            "Команди:\n"
            "/my_stats – моя статистика відвідування\n"
            "/mygroup – показати мою групу\n"
        )
        if is_admin(message.from_user.id):
            text += "\n👑 Адмін-команди:\n/adminhelp"
        bot.reply_to(message, text)
        return

    # Новий користувач – пропонуємо вибрати групу
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(
        InlineKeyboardButton("БЦІГ-25", callback_data="reg_group_БЦІГ-25"),
        InlineKeyboardButton("БЦІСТ-25 (включая ТЕ-25)", callback_data="reg_group_БЦІСТ-25")
    )
    bot.reply_to(
        message,
        "Привіт! Я бот для обліку відвідування занять 📚\n\n"
        "Спочатку зареєструйся: обери свою групу:",
        reply_markup=markup
    )

# ================== ОБРАБОТЧИК РЕГИСТРАЦИИ ==================
@bot.callback_query_handler(func=lambda call: call.data.startswith("reg_group_"))
def reg_group_callback(call):
    """Обработчик выбора группы при регистрации"""
    uid = str(call.from_user.id)
    group = call.data.split("_")[2]

    users[uid]["group"] = group
    users[uid]["awaiting_name"] = True
    save_users()

    bot.edit_message_text(
        f"Групу вибрано: {group}\n\nТепер введи своє прізвище та ім'я повністю (наприклад, Петренко Іван).",
        call.message.chat.id,
        call.message.message_id
    )

@bot.message_handler(func=lambda msg: users.get(str(msg.from_user.id), {}).get("awaiting_name"))
def process_full_name(message):
    uid = str(message.from_user.id)

    # Якщо це команда – скидаємо стан
    if message.text.startswith('/'):
        users[uid]["awaiting_name"] = False
        save_users()
        return

    full_name = message.text.strip()
    if len(full_name.split()) < 2:
        bot.reply_to(message, "Будь ласка, введи ім'я та прізвище (наприклад, Петренко Іван)")
        return

    users[uid]["full_name"] = full_name
    users[uid]["registered"] = True
    users[uid]["awaiting_name"] = False

    # Додаємо в список студентів, якщо ще немає
    existing = next((s for s in students if s.get("tg_id") == message.from_user.id), None)
    if not existing:
        students.append({
            "tg_id": message.from_user.id,
            "full_name": full_name,
            "group": users[uid]["group"],
            "username": users[uid].get("username", ""),
        })
        save_students()

    save_users()

    bot.reply_to(message, f"✅ Реєстрація завершена! Ти {full_name}, група {users[uid]['group']}.\nТепер ти можеш дивитися свою статистику через /my_stats.")

@bot.message_handler(commands=["mygroup"])
def mygroup_cmd(message):
    remember_user(message)
    uid = str(message.from_user.id)
    if users[uid].get("group"):
        bot.reply_to(message, f"📚 Твоя група: {users[uid]['group']}")
    else:
        bot.reply_to(message, "Ти ще не вибрав групу. Виконай /start для реєстрації.")

@bot.message_handler(commands=["my_stats"])
def my_stats_cmd(message):
    remember_user(message)
    uid = str(message.from_user.id)
    user = users.get(uid, {})
    if not user.get("registered"):
        bot.reply_to(message, "Спочатку зареєструйся через /start.")
        return

    student_attendance = [a for a in attendance if a.get("student_id") == message.from_user.id]

    today = date.today()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    half_year_ago = today - timedelta(days=182)

    def count_for_period(start_date):
        total = 0
        present = 0
        for rec in student_attendance:
            rec_date = datetime.strptime(rec["date"], "%Y-%m-%d").date()
            if rec_date >= start_date:
                total += 1
                if rec.get("status") == "present":
                    present += 1
        return total, present

    week_total, week_present = count_for_period(week_ago)
    month_total, month_present = count_for_period(month_ago)
    half_total, half_present = count_for_period(half_year_ago)

    response = (
        f"📊 Статистика відвідування для {user.get('full_name', '')}\n\n"
        f"За останній тиждень:\n"
        f"   Всього пар: {week_total}, був: {week_present}, пропустив: {week_total - week_present}\n\n"
        f"За останній місяць:\n"
        f"   Всього пар: {month_total}, був: {month_present}, пропустив: {month_total - month_present}\n\n"
        f"За останнє півріччя:\n"
        f"   Всього пар: {half_total}, був: {half_present}, пропустив: {half_total - half_present}\n"
    )

    bot.reply_to(message, response)

# ================== АДМИН-КОМАНДЫ ==================
@bot.message_handler(commands=["adminhelp"])
def admin_help(message):
    if not is_admin(message.from_user.id):
        return
    text = (
        "👑 Адмін-команди для обліку відвідування:\n\n"
        "/mark – відмітити присутніх на парі (починає процес)\n"
        "/add_deputy <id> – додати заступника (по id Telegram)\n"
        "/remove_deputy <id> – видалити заступника\n"
        "/list_deputies – список заступників\n"
        "/add_student – додати студента вручну (формат: ФІО, Група або ФІО, Група, ID)\n"
        "/list_students <група> – список студентів групи\n"
        "/export <група> <дата> – вивантажити журнал за день у Excel\n"
        "/export_month <група> – вивантажити статистику за останній місяць\n"
        "/export_halfyear <група> – вивантажити статистику за останнє півріччя\n"
        "/substitution – створити заміну пари на сьогодні (або на дату)\n"
        "/view_substitutions – переглянути активні заміни\n"
        "\n📋 Тимчасові заміни (діють до неділі):\n"
        "/setpair – встановити тимчасову заміну\n"
        "/resetpair – скинути конкретну заміну\n"
        "/changes – показати активні заміни\n"
        "\nПримітка: для відмітки використовуй /mark, далі вибери групу, дату, пару, і потім відмічай студентів."
    )
    bot.reply_to(message, text)

@bot.message_handler(commands=["add_deputy"])
def add_deputy(message):
    if not is_admin(message.from_user.id):
        return
    parts = message.text.split()
    if len(parts) != 2:
        bot.reply_to(message, "Формат: /add_deputy <id користувача>")
        return
    try:
        deputy_id = int(parts[1])
    except ValueError:
        bot.reply_to(message, "ID має бути числом")
        return
    ADMIN_IDS.add(deputy_id)
    bot.reply_to(message, f"✅ Користувач {deputy_id} доданий як заступник.")

@bot.message_handler(commands=["remove_deputy"])
def remove_deputy(message):
    if not is_admin(message.from_user.id):
        return
    parts = message.text.split()
    if len(parts) != 2:
        bot.reply_to(message, "Формат: /remove_deputy <id>")
        return
    try:
        deputy_id = int(parts[1])
    except ValueError:
        bot.reply_to(message, "ID має бути числом")
        return
    if deputy_id in ADMIN_IDS and deputy_id != MAIN_ADMIN_ID:
        ADMIN_IDS.remove(deputy_id)
        bot.reply_to(message, f"✅ Заступник {deputy_id} видалений.")
    else:
        bot.reply_to(message, "Не можна видалити головного адміна або неіснуючого заступника.")

@bot.message_handler(commands=["list_deputies"])
def list_deputies(message):
    if not is_admin(message.from_user.id):
        return
    deputies = list(ADMIN_IDS - {MAIN_ADMIN_ID})
    if not deputies:
        bot.reply_to(message, "Заступників поки немає.")
    else:
        text = "👥 Список заступників (ID):\n" + "\n".join(str(d) for d in deputies)
        bot.reply_to(message, text)

@bot.message_handler(commands=["add_student"])
def add_student_cmd(message):
    if not is_admin(message.from_user.id):
        return
    # Якщо є аргументи – пробуємо розпарсити одразу
    parts = message.text.split(maxsplit=1)
    if len(parts) > 1:
        # Є текст після команди
        process_add_student_direct(message, parts[1])
    else:
        # Немає аргументів – запускаємо стан очікування
        bot.reply_to(
            message,
            "Введіть дані студента у форматі:\n"
            "• Через кому: `Прізвище Ім'я, Група, ID` (ID необов'язковий)\n"
            "• Через пробіл: `Прізвище Ім'я Група ID` (ID необов'язковий)\n\n"
            "Наприклад:\n"
            "Петренко Іван, БЦІГ-25\n"
            "Петренко Іван, БЦІГ-25, 123456789\n"
            "або\n"
            "Петренко Іван БЦІГ-25 123456789",
            parse_mode="Markdown"
        )
        users[str(message.from_user.id)]["awaiting_student"] = True
        save_users()

def process_add_student_direct(message, args_text):
    """Обробка аргументів, переданих прямо в команді"""
    uid = str(message.from_user.id)
    args_text = args_text.strip()

    # Спроба 1: розділити комами
    if ',' in args_text:
        parts = [p.strip() for p in args_text.split(',')]
        if len(parts) == 2:
            full_name, group = parts
            tg_id = None
        elif len(parts) == 3:
            full_name, group, tg_id_str = parts
            try:
                tg_id = int(tg_id_str)
            except ValueError:
                bot.reply_to(message, "❌ ID має бути числом.")
                users[uid]["awaiting_student"] = False
                save_users()
                return
        else:
            bot.reply_to(message, "❌ Неправильний формат. Використовуй: ПІБ, Група, ID (або без ID)")
            users[uid]["awaiting_student"] = False
            save_users()
            return
    else:
        # Спроба 2: розділити пробілами
        tokens = args_text.split()
        if len(tokens) < 2:
            bot.reply_to(message, "❌ Неправильний формат. Використовуй: ПІБ Група або ПІБ Група ID")
            users[uid]["awaiting_student"] = False
            save_users()
            return

        # Перевіряємо, чи останній токен є числом (ID)
        last_token = tokens[-1]
        try:
            # Спробуємо перетворити останній токен на int
            tg_id = int(last_token)
            # Якщо вдалося, то ID є, і група – передостанній токен
            group = tokens[-2]
            full_name = ' '.join(tokens[:-2])
        except ValueError:
            # Останній токен не число – значить, ID немає
            tg_id = None
            group = tokens[-1]
            full_name = ' '.join(tokens[:-1])

    # Перевіряємо групу
    if group not in ["БЦІГ-25", "БЦІСТ-25"]:
        bot.reply_to(message, "❌ Група має бути БЦІГ-25 або БЦІСТ-25")
        users[uid]["awaiting_student"] = False
        save_users()
        return

    # Перевіряємо ПІБ
    if len(full_name.split()) < 2:
        bot.reply_to(message, "❌ Вкажи хоча б ім'я та прізвище")
        users[uid]["awaiting_student"] = False
        save_users()
        return

    # Додаємо студента до списку
    student_entry = {
        "tg_id": tg_id,
        "full_name": full_name,
        "group": group,
        "username": "",
    }
    students.append(student_entry)
    save_students()

    # Якщо є tg_id, створюємо/оновлюємо запис у users
    if tg_id:
        uid_str = str(tg_id)
        if uid_str not in users:
            users[uid_str] = {
                "id": tg_id,
                "username": "",
                "first_name": "",
                "role": "student",
                "group": group,
                "full_name": full_name,
                "registered": True
            }
        else:
            users[uid_str]["group"] = group
            users[uid_str]["full_name"] = full_name
            users[uid_str]["registered"] = True
        save_users()
        bot.reply_to(message, f"✅ Студента {full_name} (ID: {tg_id}) додано до групи {group} та зареєстровано в боті.")
    else:
        bot.reply_to(message, f"✅ Студента {full_name} додано до групи {group} (без ID). Коли він зайде в бот, йому треба буде зареєструватися.")

    users[uid]["awaiting_student"] = False
    save_users()

@bot.message_handler(func=lambda msg: users.get(str(msg.from_user.id), {}).get("awaiting_student"))
def process_add_student(message):
    uid = str(message.from_user.id)

    # Якщо це команда – скидаємо стан
    if message.text.startswith('/'):
        users[uid]["awaiting_student"] = False
        save_users()
        return

    # Викликаємо той самий парсер, що й для прямої передачі
    process_add_student_direct(message, message.text)

@bot.message_handler(commands=["list_students"])
def list_students_cmd(message):
    if not is_admin(message.from_user.id):
        return
    parts = message.text.split(maxsplit=1)
    if len(parts) != 2:
        bot.reply_to(message, "Формат: /list_students <група>")
        return
    group = parts[1].strip()
    group_students = get_students_by_group(group)
    if not group_students:
        bot.reply_to(message, f"У групі {group} немає студентів.")
        return
    text = f"📋 Студенти групи {group}:\n"
    for s in group_students:
        status = " (зареєстрований)" if s.get("tg_id") else " (не зареєстрований)"
        text += f"• {s['full_name']}{status}\n"
    bot.reply_to(message, text)

# ================== ОТМЕТКА ПОСЕЩАЕМОСТИ ==================
# Храним состояние процесса отметки для каждого админа
mark_states = {}

class MarkState:
    def __init__(self):
        self.step = 0          # 0: выбор группы, 1: выбор даты, 2: выбор пары, 3: отметка
        self.group = None
        self.date = None       # объект date
        self.pair_num = None   # номер пары (строкой)
        self.subject = None    # предмет (для отображения)
        self.students = []     # список студентов для отметки
        self.attendance = {}   # словарь {student_id: status} (present/absent)
        self.current_index = 0 # текущий индекс студента для отметки
        self.marked_pairs = [] # список уже отмеченных пар (номера)

@bot.message_handler(commands=["mark"])
def mark_start(message):
    if not is_admin(message.from_user.id):
        return
    mark_states[message.from_user.id] = MarkState()
    # Шаг 1: выбор группы
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(
        InlineKeyboardButton("БЦІГ-25", callback_data="mark_group_БЦІГ-25"),
        InlineKeyboardButton("БЦІСТ-25", callback_data="mark_group_БЦІСТ-25"),
        InlineKeyboardButton("❌ Скасувати", callback_data="mark_cancel")
    )
    bot.reply_to(message, "Оберіть групу для відмітки:", reply_markup=markup)

# ================== ОБРАБОТЧИКИ ДЛЯ ОТМЕТКИ ==================
# Единый обработчик для всех mark_ callback'ов
@bot.callback_query_handler(func=lambda call: call.data.startswith("mark_"))
def mark_callback_handler(call):
    """Единый обработчик для всех mark_ callback'ов"""
    admin_id = call.from_user.id
    state = mark_states.get(admin_id)

    # Обработка отмены (работает всегда)
    if call.data == "mark_cancel":
        if admin_id in mark_states:
            del mark_states[admin_id]
        bot.edit_message_text("❌ Відміну скасовано.", call.message.chat.id, call.message.message_id)
        return

    # Обработка действий после завершения пары
    if call.data in ["mark_another_pair", "mark_finish_day", "mark_finish"]:
        if not state:
            bot.answer_callback_query(call.id, "Стан не знайдено")
            return
        if call.data == "mark_another_pair":
            # Вернуться к выбору пары, исключая уже отмеченные
            choose_pair(call.message.chat.id, admin_id, call.message.message_id, exclude=state.marked_pairs)
        elif call.data == "mark_finish_day":
            # Сформировать сводный Excel за день
            export_day_report(admin_id, call.message.chat.id)
            del mark_states[admin_id]
        elif call.data == "mark_finish":
            del mark_states[admin_id]
            bot.edit_message_text("✅ Роботу завершено.", call.message.chat.id, call.message.message_id)
        return

    # Обработка выбора группы
    if call.data.startswith("mark_group_"):
        if not state:
            bot.answer_callback_query(call.id, "Спочатку почніть /mark")
            return

        state.group = call.data.split("_")[2]
        state.step = 1
        # Шаг 2: выбор даты
        markup = InlineKeyboardMarkup(row_width=3)
        today = date.today()
        buttons = [
            InlineKeyboardButton("Сьогодні", callback_data=f"mark_date_{today.isoformat()}"),
            InlineKeyboardButton("Вчора", callback_data=f"mark_date_{(today - timedelta(days=1)).isoformat()}"),
            InlineKeyboardButton("Завтра", callback_data=f"mark_date_{(today + timedelta(days=1)).isoformat()}"),
        ]
        markup.add(*buttons)
        markup.add(InlineKeyboardButton("🔢 Інша дата (введи вручну)", callback_data="mark_date_manual"))
        markup.add(InlineKeyboardButton("❌ Скасувати", callback_data="mark_cancel"))
        bot.edit_message_text(
            f"Група: {state.group}\nТепер оберіть дату:",
            call.message.chat.id,
            call.message.message_id,
            reply_markup=markup
        )
        return

    # Обработка выбора даты
    if call.data.startswith("mark_date_"):
        if not state:
            bot.answer_callback_query(call.id, "Спочатку почніть /mark")
            return

        if call.data == "mark_date_manual":
            bot.edit_message_text(
                "Введіть дату у форматі РРРР-ММ-ДД (наприклад 2025-03-20):",
                call.message.chat.id,
                call.message.message_id
            )
            state.step = 2
            return
        else:
            date_str = call.data.split("_")[2]
            state.date = datetime.strptime(date_str, "%Y-%m-%d").date()
            state.step = 2
            # Переходим к выбору пары
            choose_pair(call.message.chat.id, admin_id, call.message.message_id, exclude=state.marked_pairs)
            return

    # Обработка выбора пары
    if call.data.startswith("mark_pair_"):
        if not state:
            bot.answer_callback_query(call.id, "Спочатку почніть /mark")
            return

        pair_num = call.data.split("_")[2]
        state.pair_num = pair_num

        # Получаем расписание с учётом временных изменений и замен на дату
        day_key = get_day_key(state.date)
        week_type = get_week_type(state.date)
        group_schedule = get_schedule_with_changes(state.group, day_key, week_type)
        date_str = state.date.isoformat()
        if date_str in substitutions and state.group in substitutions[date_str] and pair_num in substitutions[date_str][state.group]:
            subj = substitutions[date_str][state.group][pair_num].get("subject", "Невідомо")
        else:
            subj = group_schedule.get(pair_num, {}).get("subject", "Невідомо")
        state.subject = subj

        # Получаем список студентов группы
        group_students = get_students_by_group(state.group)
        if not group_students:
            bot.edit_message_text("У цій групі немає студентів. Спочатку додайте через /add_student.", call.message.chat.id, call.message.message_id)
            del mark_states[admin_id]
            return

        state.students = group_students
        # Сбросить attendance для новой пары
        state.attendance = {}
        for s in group_students:
            state.attendance[s["full_name"]] = "absent"

        state.step = 4
        state.current_index = 0
        show_next_student(call.message.chat.id, admin_id, call.message.message_id)
        return

    # Обработка отметки студента
    if call.data.startswith("mark_student_"):
        if not state:
            bot.answer_callback_query(call.id, "Спочатку почніть /mark")
            return

        action = call.data.split("_")[2]
        if state.current_index >= len(state.students):
            return
        student = state.students[state.current_index]

        if action == "present":
            state.attendance[student["full_name"]] = "present"
            state.current_index += 1
        elif action == "absent":
            state.attendance[student["full_name"]] = "absent"
            state.current_index += 1
        elif action == "finish":
            # Завершить отметку текущей пары досрочно
            finish_marking(call.message.chat.id, admin_id, call.message.message_id, ask_next=True)
            return

        # Переходим к следующему студенту или завершаем
        if state.current_index < len(state.students):
            show_next_student(call.message.chat.id, admin_id, call.message.message_id)
        else:
            finish_marking(call.message.chat.id, admin_id, call.message.message_id, ask_next=True)
        return

def choose_pair(chat_id, admin_id, message_id=None, exclude=None):
    if exclude is None:
        exclude = []
    state = mark_states[admin_id]
    day_key = get_day_key(state.date)
    week_type = get_week_type(state.date)
    # Получаем расписание с временными изменениями
    group_schedule = get_schedule_with_changes(state.group, day_key, week_type)

    # Применяем замены на конкретную дату (более высокий приоритет)
    date_str = state.date.isoformat()
    if date_str in substitutions and state.group in substitutions[date_str]:
        for pair_num, sub in substitutions[date_str][state.group].items():
            group_schedule[pair_num] = sub

    pairs = []
    for p in sorted(group_schedule.keys(), key=lambda x: int(x) if x.isdigit() else 0):
        if p in exclude:
            continue
        if p == "org":
            pairs.append(("org", "Організаційна година"))
        else:
            subject = group_schedule[p].get("subject", "Невідомо")
            pairs.append((p, f"{p} пара — {subject}"))

    if not pairs:
        bot.send_message(chat_id, "Усі пари на цей день вже відмічені.")
        # Предложить завершить день
        markup = InlineKeyboardMarkup()
        markup.add(InlineKeyboardButton("📥 Завершити день і скачати звіт", callback_data="mark_finish_day"))
        markup.add(InlineKeyboardButton("❌ Завершити", callback_data="mark_finish"))
        bot.send_message(chat_id, "Що бажаєте зробити далі?", reply_markup=markup)
        return

    markup = InlineKeyboardMarkup(row_width=1)
    for p_val, p_text in pairs:
        markup.add(InlineKeyboardButton(p_text, callback_data=f"mark_pair_{p_val}"))
    markup.add(InlineKeyboardButton("❌ Скасувати", callback_data="mark_cancel"))

    if message_id:
        bot.edit_message_text(
            f"Група: {state.group}\nДата: {state.date.strftime('%d.%m.%Y')}\nОберіть пару:",
            chat_id,
            message_id,
            reply_markup=markup
        )
    else:
        bot.send_message(
            chat_id,
            f"Група: {state.group}\nДата: {state.date.strftime('%d.%m.%Y')}\nОберіть пару:",
            reply_markup=markup
        )
    state.step = 3

def show_next_student(chat_id, admin_id, message_id=None):
    state = mark_states[admin_id]

    if state.current_index >= len(state.students):
        finish_marking(chat_id, admin_id, message_id, ask_next=True)
        return

    student = state.students[state.current_index]
    full_name = student["full_name"]
    current_status = state.attendance.get(full_name, "absent")

    markup = InlineKeyboardMarkup(row_width=2)
    markup.add(
        InlineKeyboardButton("✅ Присутній", callback_data=f"mark_student_present"),
        InlineKeyboardButton("❌ Відсутній", callback_data=f"mark_student_absent")
    )
    markup.add(
        InlineKeyboardButton("⏩ Завершити", callback_data=f"mark_student_finish")
    )

    text = f"📌 {full_name}\nПоточний статус: {'Присутній' if current_status == 'present' else 'Відсутній'}\n\nОберіть статус:"

    if message_id:
        bot.edit_message_text(text, chat_id, message_id, reply_markup=markup)
    else:
        bot.send_message(chat_id, text, reply_markup=markup)

def finish_marking(chat_id, admin_id, message_id=None, ask_next=False):
    state = mark_states[admin_id]
    date_str = state.date.isoformat()

    # Сохраняем текущую пару (исправлено для студентов без ID)
    for student in state.students:
        status = state.attendance.get(student["full_name"], "absent")
        existing = None
        for idx, rec in enumerate(attendance):
            if rec.get("date") == date_str and rec.get("pair_num") == state.pair_num and rec.get("group") == state.group:
                if student.get("tg_id") is not None:
                    if rec.get("student_id") == student.get("tg_id"):
                        existing = idx
                        break
                else:
                    if rec.get("student_name") == student["full_name"]:
                        existing = idx
                        break
        if existing is not None:
            attendance[existing]["status"] = status
        else:
            attendance.append({
                "date": date_str,
                "pair_num": state.pair_num,
                "group": state.group,
                "subject": state.subject,
                "student_id": student.get("tg_id"),
                "student_name": student["full_name"],
                "status": status
            })
    save_attendance()

    # Добавляем номер пары в список отмеченных
    if state.pair_num not in state.marked_pairs:
        state.marked_pairs.append(state.pair_num)

    total = len(state.students)
    present = sum(1 for v in state.attendance.values() if v == "present")
    absent = total - present

    summary = (
        f"✅ Пара {state.pair_num} ({state.subject}) відмічена!\n"
        f"Група: {state.group}\n"
        f"Дата: {state.date.strftime('%d.%m.%Y')}\n\n"
        f"Присутні: {present}\n"
        f"Відсутні: {absent}"
    )

    if message_id:
        bot.edit_message_text(summary, chat_id, message_id)
    else:
        bot.send_message(chat_id, summary)

    if ask_next:
        # Предлагаем дальнейшие действия
        markup = InlineKeyboardMarkup(row_width=1)
        markup.add(
            InlineKeyboardButton("➕ Відмітити іншу пару", callback_data="mark_another_pair"),
            InlineKeyboardButton("📥 Завершити день і скачати звіт", callback_data="mark_finish_day"),
            InlineKeyboardButton("❌ Завершити", callback_data="mark_finish")
        )
        bot.send_message(chat_id, "Що бажаєте зробити далі?", reply_markup=markup)
    else:
        # Завершение без вопросов (например, после выбора "Завершити")
        if admin_id in mark_states:
            del mark_states[admin_id]

def export_day_report(admin_id, chat_id):
    """Сформировать отчёт за день по всем отмеченным парам (из состояния)"""
    state = mark_states.get(admin_id)
    if not state:
        bot.send_message(chat_id, "Стан не знайдено.")
        return

    date_str = state.date.isoformat()
    group = state.group
    marked_pairs = state.marked_pairs

    # Собираем все записи за этот день для этой группы по отмеченным парам
    day_records = [r for r in attendance if r.get("date") == date_str and r.get("group") == group and r.get("pair_num") in marked_pairs]
    if not day_records:
        bot.send_message(chat_id, "За цей день немає відмічених пар.")
        return

    # Получаем список уникальных студентов
    students_list = sorted(set(r["student_name"] for r in day_records))
    # Получаем список уникальных пар
    pairs = set((r["pair_num"], r.get("subject", "")) for r in day_records)
    sorted_pairs = sorted(pairs, key=lambda x: int(x[0]) if x[0] != "org" else 0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Відвідування {date_str}"

    # Заголовки
    ws.cell(row=1, column=1, value="Студент")
    for col, (pair_num, subject) in enumerate(sorted_pairs, start=2):
        ws.cell(row=1, column=col, value=f"{pair_num} ({subject})")

    # Данные
    for row, student_name in enumerate(students_list, start=2):
        ws.cell(row=row, column=1, value=student_name)
        for col, (pair_num, subject) in enumerate(sorted_pairs, start=2):
            status = next((r["status"] for r in day_records if r["student_name"] == student_name and r["pair_num"] == pair_num), None)
            ws.cell(row=row, column=col, value="Присутній" if status == "present" else ("Відсутній" if status == "absent" else ""))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    bot.send_document(chat_id, output, visible_file_name=f"attendance_{group}_{date_str}_summary.xlsx")
    bot.send_message(chat_id, "✅ Звіт за день сформовано.")

# ================== ЭКСПОРТ ЗА ПЕРИОД ==================
def generate_period_report(chat_id, group, start_date, end_date, period_name):
    """Сформировать отчёт за период (месяц/полугодие)"""
    start_str = start_date.isoformat()
    end_str = end_date.isoformat()

    records = [r for r in attendance
               if r.get("group") == group
               and r.get("date") >= start_str
               and r.get("date") <= end_str]

    if not records:
        bot.send_message(chat_id, f"За цей період немає записів для групи {group}.")
        return

    wb = openpyxl.Workbook()

    # Лист 1: Сводка по студентам
    ws_summary = wb.active
    ws_summary.title = "Зведений"

    students_list = sorted(set(r["student_name"] for r in records))
    total_pairs = len(set((r["date"], r["pair_num"]) for r in records))

    ws_summary.cell(row=1, column=1, value="Студент")
    ws_summary.cell(row=1, column=2, value="Всього пар")
    ws_summary.cell(row=1, column=3, value="Присутній")
    ws_summary.cell(row=1, column=4, value="Відсутній")
    ws_summary.cell(row=1, column=5, value="% відвідування")

    for row, student in enumerate(students_list, start=2):
        student_records = [r for r in records if r["student_name"] == student]
        present = sum(1 for r in student_records if r["status"] == "present")
        absent = len(student_records) - present
        percent = round((present / len(student_records)) * 100, 1) if student_records else 0

        ws_summary.cell(row=row, column=1, value=student)
        ws_summary.cell(row=row, column=2, value=len(student_records))
        ws_summary.cell(row=row, column=3, value=present)
        ws_summary.cell(row=row, column=4, value=absent)
        ws_summary.cell(row=row, column=5, value=f"{percent}%")

    # Лист 2: Детально
    ws_detail = wb.create_sheet("Детально")
    ws_detail.cell(row=1, column=1, value="Дата")
    ws_detail.cell(row=1, column=2, value="Пара")
    ws_detail.cell(row=1, column=3, value="Предмет")
    ws_detail.cell(row=1, column=4, value="Студент")
    ws_detail.cell(row=1, column=5, value="Статус")

    sorted_records = sorted(records, key=lambda r: (r["date"], r["pair_num"]))
    row = 2
    for r in sorted_records:
        ws_detail.cell(row=row, column=1, value=r["date"])
        ws_detail.cell(row=row, column=2, value=r["pair_num"])
        ws_detail.cell(row=row, column=3, value=r.get("subject", ""))
        ws_detail.cell(row=row, column=4, value=r["student_name"])
        ws_detail.cell(row=row, column=5, value="Присутній" if r["status"] == "present" else "Відсутній")
        row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    bot.send_document(chat_id, output, visible_file_name=f"attendance_{group}_{period_name}.xlsx")

@bot.message_handler(commands=["export_month"])
def export_month_cmd(message):
    if not is_admin(message.from_user.id):
        return
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2:
        bot.reply_to(message, "Формат: /export_month <група>")
        return
    group = parts[1]
    end_date = date.today()
    start_date = end_date - timedelta(days=30)
    generate_period_report(message.chat.id, group, start_date, end_date, "month")

@bot.message_handler(commands=["export_halfyear"])
def export_halfyear_cmd(message):
    if not is_admin(message.from_user.id):
        return
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2:
        bot.reply_to(message, "Формат: /export_halfyear <група>")
        return
    group = parts[1]
    end_date = date.today()
    start_date = end_date - timedelta(days=182)
    generate_period_report(message.chat.id, group, start_date, end_date, "halfyear")

# ================== ЗАМЕНЫ ПАР (старая система, на конкретные даты) ==================
@bot.message_handler(commands=["substitution"])
def substitution_cmd(message):
    if not is_admin(message.from_user.id):
        return
    users[str(message.from_user.id)]["subst_state"] = "awaiting_date"
    bot.reply_to(message, "Введіть дату заміни у форматі РРРР-ММ-ДД (наприклад, 2025-03-20) або 'сьогодні'/'завтра':")

@bot.message_handler(func=lambda msg: users.get(str(msg.from_user.id), {}).get("subst_state") == "awaiting_date")
def subst_date(message):
    uid = str(message.from_user.id)

    # Якщо це команда – скидаємо стан
    if message.text.startswith('/'):
        users[uid].pop("subst_state", None)
        save_users()
        return

    text = message.text.strip().lower()
    if text in ["сьогодні", "сегодня"]:
        subst_date = date.today()
    elif text in ["завтра", "tomorrow"]:
        subst_date = date.today() + timedelta(days=1)
    else:
        try:
            subst_date = datetime.strptime(text, "%Y-%m-%d").date()
        except ValueError:
            bot.reply_to(message, "Неправильний формат дати. Спробуйте ще раз або введіть 'сьогодні'/'завтра'.")
            return
    users[uid]["subst_date"] = subst_date.isoformat()
    users[uid]["subst_state"] = "awaiting_group"
    save_users()
    bot.reply_to(message, f"Дата: {subst_date.strftime('%d.%m.%Y')}\nТепер оберіть групу:", reply_markup=group_keyboard())

def group_keyboard():
    markup = InlineKeyboardMarkup(row_width=1)
    markup.add(
        InlineKeyboardButton("БЦІГ-25", callback_data="subst_group_БЦІГ-25"),
        InlineKeyboardButton("БЦІСТ-25", callback_data="subst_group_БЦІСТ-25"),
        InlineKeyboardButton("❌ Скасувати", callback_data="subst_cancel")
    )
    return markup

@bot.callback_query_handler(func=lambda call: call.data.startswith("subst_"))
def subst_callback(call):
    admin_id = call.from_user.id
    uid = str(admin_id)
    if call.data == "subst_cancel":
        users[uid].pop("subst_state", None)
        users[uid].pop("subst_date", None)
        save_users()
        bot.edit_message_text("❌ Заміну скасовано.", call.message.chat.id, call.message.message_id)
        return

    if call.data.startswith("subst_group_"):
        group = call.data.split("_")[2]
        users[uid]["subst_group"] = group
        users[uid]["subst_state"] = "awaiting_pair"
        save_users()
        bot.edit_message_text(
            f"Група: {group}\nВведіть номер пари (1-5 або 'org' для організаційної години):",
            call.message.chat.id,
            call.message.message_id
        )

@bot.message_handler(func=lambda msg: users.get(str(msg.from_user.id), {}).get("subst_state") == "awaiting_pair")
def subst_pair(message):
    uid = str(message.from_user.id)

    if message.text.startswith('/'):
        users[uid].pop("subst_state", None)
        save_users()
        return

    pair_input = message.text.strip().lower()
    if pair_input == "org":
        pair_num = "org"
    else:
        try:
            pair_num = str(int(pair_input))
            if int(pair_num) < 1 or int(pair_num) > 5:
                raise ValueError
        except ValueError:
            bot.reply_to(message, "Номер пари має бути від 1 до 5 або 'org'")
            return
    users[uid]["subst_pair"] = pair_num
    users[uid]["subst_state"] = "awaiting_subject"
    save_users()
    bot.reply_to(message, "Введіть новий предмет (або пропустіть, якщо не змінюється):")

@bot.message_handler(func=lambda msg: users.get(str(msg.from_user.id), {}).get("subst_state") == "awaiting_subject")
def subst_subject(message):
    uid = str(message.from_user.id)

    if message.text.startswith('/'):
        users[uid].pop("subst_state", None)
        save_users()
        return

    subject = message.text.strip()
    users[uid]["subst_subject"] = subject if subject else None
    users[uid]["subst_state"] = "awaiting_room"
    save_users()
    bot.reply_to(message, "Введіть нову аудиторію (або пропустіть):")

@bot.message_handler(func=lambda msg: users.get(str(msg.from_user.id), {}).get("subst_state") == "awaiting_room")
def subst_room(message):
    uid = str(message.from_user.id)

    if message.text.startswith('/'):
        users[uid].pop("subst_state", None)
        save_users()
        return

    room = message.text.strip()
    users[uid]["subst_room"] = room if room else None
    users[uid]["subst_state"] = "awaiting_teacher"
    save_users()
    bot.reply_to(message, "Введіть нового викладача (або пропустіть):")

@bot.message_handler(func=lambda msg: users.get(str(msg.from_user.id), {}).get("subst_state") == "awaiting_teacher")
def subst_teacher(message):
    uid = str(message.from_user.id)

    if message.text.startswith('/'):
        users[uid].pop("subst_state", None)
        save_users()
        return

    teacher = message.text.strip()
    users[uid]["subst_teacher"] = teacher if teacher else None

    date_str = users[uid]["subst_date"]
    group = users[uid]["subst_group"]
    pair_num = users[uid]["subst_pair"]

    if date_str not in substitutions:
        substitutions[date_str] = {}
    if group not in substitutions[date_str]:
        substitutions[date_str][group] = {}

    day_key = get_day_key(datetime.strptime(date_str, "%Y-%m-%d").date())
    week_type = get_week_type(datetime.strptime(date_str, "%Y-%m-%d").date())
    original = schedule.get(group, {}).get(day_key, {}).get(week_type, {}).get(pair_num, {})

    sub_entry = {
        "subject": users[uid]["subst_subject"] if users[uid]["subst_subject"] else original.get("subject", ""),
        "room": users[uid]["subst_room"] if users[uid]["subst_room"] else original.get("room", ""),
        "teacher": users[uid]["subst_teacher"] if users[uid]["subst_teacher"] else original.get("teacher", "")
    }
    substitutions[date_str][group][pair_num] = sub_entry
    save_substitutions()

    bot.reply_to(message, f"✅ Заміну для {date_str}, група {group}, пара {pair_num} збережено.")

    del users[uid]["subst_state"]
    for k in ["subst_date", "subst_group", "subst_pair", "subst_subject", "subst_room", "subst_teacher"]:
        users[uid].pop(k, None)
    save_users()

@bot.message_handler(commands=["view_substitutions"])
def view_substitutions(message):
    if not is_admin(message.from_user.id):
        return
    if not substitutions:
        bot.reply_to(message, "Немає активних замін.")
        return
    text = "📅 Активні заміни:\n"
    for date_str, groups in substitutions.items():
        text += f"\n{date_str}:\n"
        for group, pairs in groups.items():
            for pair_num, sub in pairs.items():
                text += f"  {group}, пара {pair_num}: {sub.get('subject')} ({sub.get('room')}) – {sub.get('teacher')}\n"
    bot.reply_to(message, text)

# ================== НОВАЯ СИСТЕМА ВРЕМЕННЫХ ИЗМЕНЕНИЙ (команды) ==================
@bot.message_handler(commands=["setpair"])
def setpair_cmd(message):
    remember_user(message)
    if not is_admin(message):
        return

    try:
        _, rest = message.text.split(" ", 1)
    except ValueError:
        bot.reply_to(message,
            "Формат: /setpair <група> <день> <номер> <тиждень> <предмет> ; <аудиторія> ; <викладач>\n"
            "Примеры:\n"
            "/setpair БЦІГ-25 понеділок 1 чисельник Фізика ; 129 ; Гуленко І.А.\n"
            "/setpair БЦІСТ-25 середа 2 знаменник Математика ; 121 ; Приймак О.В.\n"
            "📝 Изменение действует до конца недели (воскресенья)"
        )
        return

    parts = rest.split(maxsplit=10)
    if len(parts) < 6:
        bot.reply_to(message, "Недостатньо параметрів")
        return

    group_name, day_raw, pair_str, week_raw, subject_rest = parts[0], parts[1], parts[2], parts[3], parts[4]

    if group_name not in ["БЦІГ-25", "БЦІСТ-25"]:
        bot.reply_to(message, 
            f"Невірна група. Доступні групи:\n"
            f"• БЦІГ-25\n"
            f"• БЦІСТ-25 (включая ТЕ-25)"
        )
        return

    day_key = DAY_ALIASES.get(day_raw.lower())
    if not day_key:
        bot.reply_to(message, "Невірний день")
        return

    try:
        pair_num = int(pair_str)
        if pair_num < 1 or pair_num > 6:
            bot.reply_to(message, "Номер пари повинен бути від 1 до 6")
            return
    except ValueError:
        bot.reply_to(message, "Номер пари має бути числом")
        return

    w_raw = week_raw.lower()
    if w_raw.startswith("чис"):
        week_type = "чисельник"
    elif w_raw.startswith("зн"):
        week_type = "знаменник"
    else:
        bot.reply_to(message, "Невірний тип тижня")
        return

    if ";" in subject_rest:
        parts2 = subject_rest.split(";", 2)
        subject = parts2[0].strip()
        room = parts2[1].strip() if len(parts2) > 1 else ""
        teacher = parts2[2].strip() if len(parts2) > 2 else ""
    else:
        subject = subject_rest.strip()
        room = ""
        teacher = ""

    if group_name not in temp_changes:
        temp_changes[group_name] = {}

    if day_key not in temp_changes[group_name]:
        temp_changes[group_name][day_key] = {}

    if week_type not in temp_changes[group_name][day_key]:
        temp_changes[group_name][day_key][week_type] = {}

    temp_changes[group_name][day_key][week_type][str(pair_num)] = {
        "subject": subject,
        "room": room,
        "teacher": teacher,
        "changed_at": (datetime.utcnow() + timedelta(hours=2)).strftime("%Y-%m-%d %H:%M:%S"),
        "changed_by": message.from_user.id,
        "original_subject": schedule[group_name].get(day_key, {}).get(week_type, {}).get(str(pair_num), {}).get("subject", "")
    }

    save_temp_changes()

    now_local = datetime.utcnow() + timedelta(hours=2)
    record = {
        "timestamp": now_local.strftime("%Y-%m-%d %H:%M:%S"),
        "group": group_name,
        "day_key": day_key,
        "pair_num": pair_num,
        "week_type": week_type,
        "subject": subject,
        "room": room,
        "teacher": teacher,
        "admin_id": message.from_user.id,
        "admin_username": message.from_user.username or "",
        "admin_first_name": message.from_user.first_name or "",
        "change_type": "temporary"
    }
    changelog.append(record)
    save_changelog()

    time_txt = get_pair_time(day_key, pair_num) or "час ?"
    bot.reply_to(
        message,
        f"✅ Встановлено тимчасову заміну для групи {group_name} (діє до неділі):\n"
        f"{DAYS_RU[day_key]}, пара {pair_num} ({week_type})\n"
        f"{time_txt} — {subject} {f'({room})' if room else ''} {f'— {teacher}' if teacher else ''}"
    )

@bot.message_handler(commands=["resetpair"])
def resetpair_cmd(message):
    remember_user(message)
    if not is_admin(message):
        return

    try:
        _, rest = message.text.split(" ", 1)
    except ValueError:
        bot.reply_to(message,
            "Формат: /resetpair <група> <день> <номер> <тиждень>\n"
            "Примеры:\n"
            "/resetpair БЦІГ-25 понеділок 1 чисельник\n"
            "/resetpair БЦІСТ-25 середа 2 знаменник\n"
            "Скине тимчасову заміну для вказаної пари"
        )
        return

    parts = rest.split(maxsplit=4)
    if len(parts) < 4:
        bot.reply_to(message, "Недостатньо параметрів")
        return

    group_name, day_raw, pair_str, week_raw = parts[0], parts[1], parts[2], parts[3]

    if group_name not in ["БЦІГ-25", "БЦІСТ-25"]:
        bot.reply_to(message, f"Невірна група. Використовуйте: БЦІГ-25 або БЦІСТ-25")
        return

    day_key = DAY_ALIASES.get(day_raw.lower())
    if not day_key:
        bot.reply_to(message, "Невірний день")
        return

    try:
        pair_num = int(pair_str)
    except ValueError:
        bot.reply_to(message, "Номер пари має бути числом")
        return

    w_raw = week_raw.lower()
    if w_raw.startswith("чис"):
        week_type = "чисельник"
    elif w_raw.startswith("зн"):
        week_type = "знаменник"
    else:
        bot.reply_to(message, "Невірний тип тижня")
        return

    if (group_name in temp_changes and 
        day_key in temp_changes[group_name] and 
        week_type in temp_changes[group_name][day_key] and
        str(pair_num) in temp_changes[group_name][day_key][week_type]):

        del temp_changes[group_name][day_key][week_type][str(pair_num)]

        if not temp_changes[group_name][day_key][week_type]:
            del temp_changes[group_name][day_key][week_type]
        if not temp_changes[group_name][day_key]:
            del temp_changes[group_name][day_key]
        if not temp_changes[group_name]:
            del temp_changes[group_name]

        save_temp_changes()

        now_local = datetime.utcnow() + timedelta(hours=2)
        record = {
            "timestamp": now_local.strftime("%Y-%m-%d %H:%M:%S"),
            "group": group_name,
            "day_key": day_key,
            "pair_num": pair_num,
            "week_type": week_type,
            "action": "reset_temporary_change",
            "admin_id": message.from_user.id,
            "admin_username": message.from_user.username or "",
            "admin_first_name": message.from_user.first_name or "",
        }
        changelog.append(record)
        save_changelog()

        bot.reply_to(message, f"✅ Тимчасову заміну для групи {group_name}, {DAYS_RU[day_key]}, пара {pair_num} ({week_type}) скинуто")
    else:
        bot.reply_to(message, f"Тимчасової заміни для вказаної пари не знайдено")

@bot.message_handler(commands=["changes"])
def changes_cmd(message):
    remember_user(message)
    if not is_admin(message):
        return

    has_changes_bcig = any(temp_changes.get("БЦІГ-25", {}).values())
    has_changes_bcist = any(temp_changes.get("БЦІСТ-25", {}).values())

    if not has_changes_bcig and not has_changes_bcist:
        bot.reply_to(message, "📋 Активних тимчасових замін немає для жодної групи.")
        return

    lines = ["📋 Активні тимчасові заміни (діють до неділі):\n"]

    for group_name in ["БЦІГ-25", "БЦІСТ-25"]:
        if group_name in temp_changes and temp_changes[group_name]:
            lines.append(f"\n👥 Група: {group_name}")
            for day_key, day_data in temp_changes[group_name].items():
                lines.append(f"  📅 {DAYS_RU.get(day_key, day_key)}")
                for week_type, week_data in day_data.items():
                    if week_data:
                        lines.append(f"    🔹 {week_type.upper()}:")
                        for pair_num, change in week_data.items():
                            subject = change.get("subject", "—")
                            room = change.get("room", "")
                            teacher = change.get("teacher", "")
                            changed_at = change.get("changed_at", "")
                            original = change.get("original_subject", "")
                            original_info = f" (було: {original})" if original else ""
                            line = f"      {pair_num}) {subject}{original_info}"
                            if room:
                                line += f" ({room})"
                            if teacher:
                                line += f" — {teacher}"
                            if changed_at:
                                try:
                                    dt = datetime.strptime(changed_at, "%Y-%m-%d %H:%M:%S")
                                    line += f" | змінено: {dt.strftime('%d.%m %H:%M')}"
                                except:
                                    pass
                            lines.append(line)

    text = "\n".join(lines)
    if len(text) > 4000:
        for i in range(0, len(text), 4000):
            bot.reply_to(message, text[i:i + 4000])
    else:
        bot.reply_to(message, text)

# ================== ВЫГРУЗКА ЗА ДЕНЬ (старая команда) ==================
@bot.message_handler(commands=["export"])
def export_cmd(message):
    if not is_admin(message.from_user.id):
        return
    parts = message.text.split(maxsplit=2)
    if len(parts) < 3:
        bot.reply_to(message, "Формат: /export <група> <дата РРРР-ММ-ДД>")
        return
    group = parts[1]
    date_str = parts[2]
    try:
        export_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        bot.reply_to(message, "Неправильний формат дати. Використовуйте РРРР-ММ-ДД")
        return

    day_records = [r for r in attendance if r.get("date") == date_str and r.get("group") == group]
    if not day_records:
        bot.reply_to(message, "За цей день немає записів про відвідування.")
        return

    pairs = set((r["pair_num"], r.get("subject", "")) for r in day_records)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Відвідування {date_str}"

    students_list = sorted(set(r["student_name"] for r in day_records))
    sorted_pairs = sorted(pairs, key=lambda x: int(x[0]) if x[0] != "org" else 0)

    ws.cell(row=1, column=1, value="Студент")
    for col, (pair_num, subject) in enumerate(sorted_pairs, start=2):
        ws.cell(row=1, column=col, value=f"{pair_num} ({subject})")

    for row, student_name in enumerate(students_list, start=2):
        ws.cell(row=row, column=1, value=student_name)
        for col, (pair_num, subject) in enumerate(sorted_pairs, start=2):
            status = next((r["status"] for r in day_records if r["student_name"] == student_name and r["pair_num"] == pair_num), None)
            ws.cell(row=row, column=col, value="Присутній" if status == "present" else ("Відсутній" if status == "absent" else ""))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    bot.send_document(message.chat.id, output, visible_file_name=f"attendance_{group}_{date_str}.xlsx")

# ================== АВТОМАТИЧЕСКИЙ СБРОС ВРЕМЕННЫХ ИЗМЕНЕНИЙ ==================
def auto_reset_temp_changes():
    """Автоматически сбрасывает временные изменения в воскресенье в 23:00"""
    while True:
        try:
            now = datetime.utcnow() + timedelta(hours=2)
            if now.weekday() == 6 and now.hour == 23 and now.minute == 0:
                print(f"[{now.strftime('%Y-%m-%d %H:%M')}] Автоматичне скидання тимчасових замін...")
                changed_groups = []
                for group_name in ["БЦІГ-25", "БЦІСТ-25"]:
                    if group_name in temp_changes and temp_changes[group_name]:
                        temp_changes[group_name] = {}
                        changed_groups.append(group_name)
                        print(f"✅ Скинуті тимчасові заміни для {group_name}")
                if changed_groups:
                    save_temp_changes()
                    for admin_id in ADMIN_IDS:
                        try:
                            bot.send_message(
                                admin_id,
                                "🔄 Автоматичне оновлення розкладу:\n"
                                "✅ Всі тимчасові заміни скинуті.\n"
                                f"Групи: {', '.join(changed_groups)}\n"
                                "Розклад на наступний тиждень повернуто до стандартного."
                            )
                        except Exception as e:
                            print(f"Не вдалося відправити сповіщення адміну {admin_id}: {e}")
                else:
                    print("Немає тимчасових замін для скидання")
                time.sleep(24 * 3600)
            else:
                time.sleep(60)
        except Exception as e:
            print(f"Помилка в auto_reset_temp_changes: {e}")
            time.sleep(300)

threading.Thread(target=auto_reset_temp_changes, daemon=True).start()

# ================== ЗАПУСК БОТА ==================
if __name__ == "__main__":
    print("Attendance bot started")
    print(f"Головний адмін: {MAIN_ADMIN_ID}")
    print(f"Адміни: {ADMIN_IDS}")
    print("✅ Система тимчасових замін активна (команди /setpair, /resetpair, /changes)")
    print("🔄 Автоматичне скидання тимчасових замін: щонеділі о 23:00")
    bot.infinity_polling()
